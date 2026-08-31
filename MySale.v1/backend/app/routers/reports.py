from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Optional
from datetime import date, datetime, timedelta
from io import BytesIO
from app.database import get_db
from app.models.user import User, RoleType
from app.models.sale import Sale, SaleItem, PaymentMethod
from app.models.shift import Shift
from app.models.inventory import Product, ProductStock, Group, Family, SubFamily, StockMovement, MovementType
from app.models.location import Location
from app.models.loss import Loss
from app.models.expense import Expense
from app.schemas.reports import (
    SalesReportRequest, SalesReportResponse, SaleDetailReport,
    InventoryReportResponse, ProductStockReport,
    EmployeeReportResponse, EmployeeShiftReport,
    EmployeesReportResponse, EmployeeSummary,
    ProfitabilityReportResponse, ProfitabilitySummary, ProfitabilityByDay,
    PurchasesReportResponse, PurchaseDetail
)
from app.utils.auth import get_current_user, require_role
from app.utils.location_scope import scoped_location_id

router = APIRouter(prefix="/api/reports", tags=["Reportes"])


def _get_tenant_location_ids(db: Session, current_user: User) -> list:
    """Sedes que el usuario puede reportar: su sede fija, o las de su negocio."""
    if current_user.location_id:
        return [current_user.location_id]
    if not current_user.tenant_id:
        return []
    locs = db.query(Location.id).filter(Location.tenant_id == current_user.tenant_id).all()
    return [l[0] for l in locs]


@router.post("/sales", response_model=SalesReportResponse)
async def get_sales_report(
    request: SalesReportRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(RoleType.SUPERUSER, RoleType.ADMIN))
):
    location_id = scoped_location_id(current_user, request.location_id)
    tenant_loc_ids = _get_tenant_location_ids(db, current_user)
    query = db.query(Sale).filter(
        Sale.created_at >= datetime.combine(request.start_date, datetime.min.time()),
        Sale.created_at <= datetime.combine(request.end_date, datetime.max.time())
    )
    if tenant_loc_ids:
        query = query.filter(Sale.location_id.in_(tenant_loc_ids))
    
    if location_id:
        query = query.filter(Sale.location_id == location_id)
    if request.cashier_id:
        query = query.filter(Sale.cashier_id == request.cashier_id)
    
    sales = query.all()
    
    total_sales = sum(s.total for s in sales)
    total_cash = sum(s.total for s in sales if s.payment_method == PaymentMethod.CASH)
    total_card = sum(s.total for s in sales if s.payment_method == PaymentMethod.CARD)
    total_transfer = sum(s.total for s in sales if s.payment_method == PaymentMethod.TRANSFER)
    total_nequi = sum(s.total for s in sales if s.payment_method == PaymentMethod.NEQUI)
    total_breb = sum(s.total for s in sales if s.payment_method == PaymentMethod.BREB)
    
    sales_by_day = {}
    for sale in sales:
        day = sale.created_at.date().isoformat()
        if day not in sales_by_day:
            sales_by_day[day] = {"date": day, "total": 0, "count": 0}
        sales_by_day[day]["total"] += sale.total
        sales_by_day[day]["count"] += 1
    
    details = []
    for sale in sales:
        location = db.query(Location).filter(Location.id == sale.location_id).first()
        cashier = db.query(User).filter(User.id == sale.cashier_id).first()
        
        for item in sale.items:
            product = db.query(Product).filter(Product.id == item.product_id).first()
            details.append(SaleDetailReport(
                sale_id=sale.id,
                folio=sale.folio,
                date=sale.created_at,
                time=sale.created_at.strftime("%H:%M:%S"),
                product_name=product.name if product else "N/A",
                product_code=product.code if product else "N/A",
                quantity=item.quantity,
                unit_price=item.unit_price,
                subtotal=item.subtotal,
                payment_method=sale.payment_method,
                cashier_name=cashier.full_name if cashier else "N/A",
                location_name=location.name if location else "N/A"
            ))
    
    return SalesReportResponse(
        total_sales=total_sales,
        total_transactions=len(sales),
        total_cash=total_cash,
        total_card=total_card,
        total_transfer=total_transfer,
        total_nequi=total_nequi,
        total_breb=total_breb,
        sales_by_day=list(sales_by_day.values()),
        details=details
    )


@router.get("/inventory", response_model=InventoryReportResponse)
async def get_inventory_report(
    location_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(RoleType.SUPERUSER, RoleType.ADMIN))
):
    location_id = scoped_location_id(current_user, location_id)
    tenant_loc_ids = _get_tenant_location_ids(db, current_user)
    location = None
    if location_id:
        location = db.query(Location).filter(Location.id == location_id).first()
        if not location:
            raise HTTPException(status_code=404, detail="Ubicacion no encontrada")
    
    query = db.query(ProductStock).join(Product)
    if tenant_loc_ids:
        query = query.filter(ProductStock.location_id.in_(tenant_loc_ids))
    if location_id:
        query = query.filter(ProductStock.location_id == location_id)
    
    stocks = query.all()
    
    products = []
    total_value = 0.0
    low_stock_count = 0
    high_stock_count = 0
    
    for stock in stocks:
        product = stock.product
        subfamily = db.query(SubFamily).filter(SubFamily.id == product.subfamily_id).first()
        family = db.query(Family).filter(Family.id == subfamily.family_id).first() if subfamily else None
        group = db.query(Group).filter(Group.id == family.group_id).first() if family else None
        
        stock_value = stock.quantity * product.weighted_cost
        total_value += stock_value
        
        status = "normal"
        if stock.quantity <= product.min_stock:
            status = "low"
            low_stock_count += 1
        elif stock.quantity >= product.max_stock:
            status = "high"
            high_stock_count += 1
        
        products.append(ProductStockReport(
            product_id=product.id,
            product_code=product.code,
            product_name=product.name,
            group_name=group.name if group else "N/A",
            family_name=family.name if family else "N/A",
            subfamily_name=subfamily.name if subfamily else "N/A",
            unit=product.unit,
            sale_price=product.sale_price,
            weighted_cost=product.weighted_cost,
            quantity=stock.quantity,
            min_stock=product.min_stock,
            max_stock=product.max_stock,
            stock_value=stock_value,
            status=status
        ))
    
    return InventoryReportResponse(
        location_id=location_id,
        location_name=location.name if location else "Todas las ubicaciones",
        total_products=len(products),
        total_stock_value=total_value,
        low_stock_count=low_stock_count,
        high_stock_count=high_stock_count,
        products=products
    )


@router.get("/employees/{user_id}", response_model=EmployeeReportResponse)
async def get_employee_report(
    user_id: int,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(RoleType.SUPERUSER, RoleType.ADMIN))
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    if current_user.tenant_id and user.tenant_id != current_user.tenant_id:
        raise HTTPException(status_code=403, detail="No tienes acceso a este empleado")
    
    if not start_date:
        start_date = date.today() - timedelta(days=30)
    if not end_date:
        end_date = date.today()
    
    shifts = db.query(Shift).filter(
        Shift.user_id == user_id,
        Shift.start_time >= datetime.combine(start_date, datetime.min.time()),
        Shift.start_time <= datetime.combine(end_date, datetime.max.time())
    ).all()
    
    total_hours = 0.0
    total_sales = 0.0
    total_transactions = 0
    shift_reports = []
    
    for shift in shifts:
        hours = 0.0
        if shift.end_time:
            delta = shift.end_time - shift.start_time
            hours = delta.total_seconds() / 3600
        
        total_hours += hours
        total_sales += shift.total_sales
        
        sales_count = db.query(Sale).filter(Sale.shift_id == shift.id).count()
        total_transactions += sales_count
        
        shift_reports.append(EmployeeShiftReport(
            shift_id=shift.id,
            date=shift.start_time.date(),
            start_time=shift.start_time,
            end_time=shift.end_time,
            hours_worked=hours if shift.end_time else None,
            total_sales=shift.total_sales,
            transactions_count=sales_count
        ))
    
    from app.models.shift import ShiftAlert
    alerts_count = db.query(ShiftAlert).filter(ShiftAlert.user_id == user_id).count()
    
    return EmployeeReportResponse(
        user_id=user.id,
        user_name=user.full_name,
        total_hours=total_hours,
        total_sales=total_sales,
        total_transactions=total_transactions,
        points=user.points,
        alerts_count=alerts_count,
        shifts=shift_reports
    )


@router.get("/export/sales/excel")
async def export_sales_excel(
    start_date: date,
    end_date: date,
    location_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(RoleType.SUPERUSER, RoleType.ADMIN))
):
    from openpyxl import Workbook
    
    location_id = scoped_location_id(current_user, location_id)
    tenant_loc_ids = _get_tenant_location_ids(db, current_user)
    query = db.query(Sale).filter(
        Sale.created_at >= datetime.combine(start_date, datetime.min.time()),
        Sale.created_at <= datetime.combine(end_date, datetime.max.time())
    )
    if tenant_loc_ids:
        query = query.filter(Sale.location_id.in_(tenant_loc_ids))
    
    if location_id:
        query = query.filter(Sale.location_id == location_id)
    
    sales = query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    
    headers = ["Folio", "Fecha", "Hora", "Producto", "Codigo", "Cantidad", 
               "Precio Unitario", "Subtotal", "Metodo de Pago", "Cajero", "Ubicacion"]
    ws.append(headers)
    
    for sale in sales:
        location = db.query(Location).filter(Location.id == sale.location_id).first()
        cashier = db.query(User).filter(User.id == sale.cashier_id).first()
        
        for item in sale.items:
            product = db.query(Product).filter(Product.id == item.product_id).first()
            ws.append([
                sale.folio,
                sale.created_at.strftime("%Y-%m-%d"),
                sale.created_at.strftime("%H:%M:%S"),
                product.name if product else "N/A",
                product.code if product else "N/A",
                item.quantity,
                item.unit_price,
                item.subtotal,
                sale.payment_method.value,
                cashier.full_name if cashier else "N/A",
                location.name if location else "N/A"
            ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"ventas_{start_date}_{end_date}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/inventory/excel")
async def export_inventory_excel(
    location_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(RoleType.SUPERUSER, RoleType.ADMIN))
):
    from openpyxl import Workbook
    
    location_id = scoped_location_id(current_user, location_id)
    tenant_loc_ids = _get_tenant_location_ids(db, current_user)
    query = db.query(ProductStock).join(Product)
    if tenant_loc_ids:
        query = query.filter(ProductStock.location_id.in_(tenant_loc_ids))
    if location_id:
        query = query.filter(ProductStock.location_id == location_id)
    
    stocks = query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    headers = ["Codigo", "Producto", "Grupo", "Familia", "Subfamilia", "Unidad",
               "Precio Venta", "Costo Ponderado", "Cantidad", "Stock Min", "Stock Max",
               "Valor Stock", "Estado"]
    ws.append(headers)
    
    for stock in stocks:
        product = stock.product
        subfamily = db.query(SubFamily).filter(SubFamily.id == product.subfamily_id).first()
        family = db.query(Family).filter(Family.id == subfamily.family_id).first() if subfamily else None
        group = db.query(Group).filter(Group.id == family.group_id).first() if family else None
        
        stock_value = stock.quantity * product.weighted_cost
        
        status = "Normal"
        if stock.quantity <= product.min_stock:
            status = "Bajo"
        elif stock.quantity >= product.max_stock:
            status = "Alto"
        
        ws.append([
            product.code,
            product.name,
            group.name if group else "N/A",
            family.name if family else "N/A",
            subfamily.name if subfamily else "N/A",
            product.unit,
            product.sale_price,
            product.weighted_cost,
            stock.quantity,
            product.min_stock,
            product.max_stock,
            stock_value,
            status
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    location = db.query(Location).filter(Location.id == location_id).first() if location_id else None
    loc_name = location.code if location else "todas"
    filename = f"inventario_{loc_name}_{date.today()}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/employees/excel")
async def export_employees_excel(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    user_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(RoleType.SUPERUSER, RoleType.ADMIN))
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    if not start_date:
        start_date = date.today() - timedelta(days=30)
    if not end_date:
        end_date = date.today()
    
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())
    
    query = db.query(User)
    if current_user.tenant_id:
        query = query.filter(User.tenant_id == current_user.tenant_id)
    if user_id:
        query = query.filter(User.id == user_id)
    users = query.all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Empleados"
    
    # Title
    ws.append([f"Reporte de Empleados - {start_date} a {end_date}"])
    ws.append([])
    
    headers = ["#", "Empleado", "Usuario", "Rol", "Turnos", "Horas Totales",
               "Prom. Hrs/Turno", "Ventas Totales", "Transacciones",
               "Prom. Venta/Turno", "Puntos", "Estado"]
    ws.append(headers)
    
    # Bold headers
    for col in range(1, len(headers) + 1):
        ws.cell(row=3, column=col).font = Font(bold=True)
    
    row_num = 0
    for user in users:
        shifts = db.query(Shift).filter(
            Shift.user_id == user.id,
            Shift.start_time >= start_dt,
            Shift.start_time <= end_dt
        ).all()
        
        total_hours = 0.0
        total_sales = 0.0
        total_transactions = 0
        
        for shift in shifts:
            if shift.end_time:
                delta = shift.end_time - shift.start_time
                total_hours += delta.total_seconds() / 3600
            total_sales += shift.total_sales
            sales_count = db.query(Sale).filter(Sale.shift_id == shift.id).count()
            total_transactions += sales_count
        
        num_shifts = len(shifts)
        role_name = user.role.name if user.role else "Sin rol"
        avg_sales = round(total_sales / num_shifts, 2) if num_shifts > 0 else 0
        avg_hours = round(total_hours / num_shifts, 2) if num_shifts > 0 else 0
        
        row_num += 1
        ws.append([
            row_num,
            user.full_name,
            user.username,
            role_name,
            num_shifts,
            round(total_hours, 2),
            avg_hours,
            total_sales,
            total_transactions,
            avg_sales,
            user.points,
            "Activo" if user.is_active else "Inactivo"
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"empleados_{start_date}_{end_date}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/profitability/excel")
async def export_profitability_excel(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    location_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(RoleType.SUPERUSER, RoleType.ADMIN))
):
    from openpyxl import Workbook
    from openpyxl.styles import Font
    
    location_id = scoped_location_id(current_user, location_id)
    tenant_loc_ids = _get_tenant_location_ids(db, current_user)
    
    if not start_date:
        start_date = date.today() - timedelta(days=30)
    if not end_date:
        end_date = date.today()
    
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())
    
    location_name = "Todas las ubicaciones"
    if location_id:
        location = db.query(Location).filter(Location.id == location_id).first()
        location_name = location.name if location else "Todas las ubicaciones"
    
    # Sales
    sales_query = db.query(Sale).filter(Sale.created_at >= start_dt, Sale.created_at <= end_dt)
    if tenant_loc_ids:
        sales_query = sales_query.filter(Sale.location_id.in_(tenant_loc_ids))
    if location_id:
        sales_query = sales_query.filter(Sale.location_id == location_id)
    sales = sales_query.all()
    total_sales = sum(s.total for s in sales)
    
    # COGS
    total_cogs = 0.0
    for sale in sales:
        for item in sale.items:
            product = db.query(Product).filter(Product.id == item.product_id).first()
            if product:
                total_cogs += item.quantity * product.weighted_cost
    
    # Expenses
    expenses_query = db.query(Expense).filter(Expense.expense_date >= start_dt, Expense.expense_date <= end_dt)
    if tenant_loc_ids:
        expenses_query = expenses_query.filter(Expense.location_id.in_(tenant_loc_ids))
    if location_id:
        expenses_query = expenses_query.filter(Expense.location_id == location_id)
    expenses = expenses_query.all()
    total_expenses = sum(e.amount for e in expenses)
    
    expenses_by_cat = {}
    for e in expenses:
        cat = e.category.value if hasattr(e.category, 'value') else str(e.category)
        cat_labels = {'purchase': 'Compras', 'utilities': 'Servicios', 'rent': 'Arriendo',
                      'salary': 'Salarios', 'maintenance': 'Mantenimiento', 'supplies': 'Insumos', 'other': 'Otros'}
        label = cat_labels.get(cat, cat)
        expenses_by_cat[label] = expenses_by_cat.get(label, 0) + e.amount
    
    # Losses
    losses_query = db.query(Loss).filter(Loss.created_at >= start_dt, Loss.created_at <= end_dt)
    if tenant_loc_ids:
        losses_query = losses_query.filter(Loss.location_id.in_(tenant_loc_ids))
    if location_id:
        losses_query = losses_query.filter(Loss.location_id == location_id)
    losses = losses_query.all()
    total_losses = sum(l.total_value for l in losses)
    
    losses_by_type = {}
    for l in losses:
        lt = l.loss_type.value if hasattr(l.loss_type, 'value') else str(l.loss_type)
        type_labels = {'breakage': 'Rotura', 'expiration': 'Vencimiento', 'theft': 'Robo',
                       'damage': 'Dano', 'other': 'Otros'}
        label = type_labels.get(lt, lt)
        losses_by_type[label] = losses_by_type.get(label, 0) + l.total_value
    
    gross_profit = total_sales - total_cogs
    net_profit = gross_profit - total_expenses - total_losses
    gross_margin = (gross_profit / total_sales * 100) if total_sales > 0 else 0
    net_margin = (net_profit / total_sales * 100) if total_sales > 0 else 0
    
    wb = Workbook()
    
    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.append([f"Reporte de Rentabilidad - {start_date} a {end_date}"])
    ws1.append([f"Ubicacion: {location_name}"])
    ws1.append([])
    
    ws1.append(["Concepto", "Monto", "% sobre Ventas"])
    for col in range(1, 4):
        ws1.cell(row=4, column=col).font = Font(bold=True)
    
    ws1.append(["Ventas Totales", total_sales, "100%"])
    ws1.append(["(-) Costo de Productos", round(total_cogs, 2), f"{round(total_cogs/total_sales*100, 1) if total_sales > 0 else 0}%"])
    ws1.append(["= Utilidad Bruta", round(gross_profit, 2), f"{round(gross_margin, 1)}%"])
    ws1.append(["(-) Gastos Operacionales", total_expenses, f"{round(total_expenses/total_sales*100, 1) if total_sales > 0 else 0}%"])
    
    for cat, amount in expenses_by_cat.items():
        ws1.append([f"    {cat}", amount, f"{round(amount/total_sales*100, 1) if total_sales > 0 else 0}%"])
    
    ws1.append(["(-) Mermas / Perdidas", total_losses, f"{round(total_losses/total_sales*100, 1) if total_sales > 0 else 0}%"])
    
    for lt, amount in losses_by_type.items():
        ws1.append([f"    {lt}", amount, f"{round(amount/total_sales*100, 1) if total_sales > 0 else 0}%"])
    
    ws1.append([])
    ws1.append(["= UTILIDAD NETA", round(net_profit, 2), f"{round(net_margin, 1)}%"])
    last_row = ws1.max_row
    for col in range(1, 4):
        ws1.cell(row=last_row, column=col).font = Font(bold=True)
    
    # Sheet 2: Daily detail
    ws2 = wb.create_sheet("Detalle Diario")
    ws2.append(["Fecha", "Ventas", "Costo Prod.", "Gastos", "Mermas", "Utilidad Bruta", "Utilidad Neta"])
    for col in range(1, 8):
        ws2.cell(row=1, column=col).font = Font(bold=True)
    
    by_day_data = {}
    for sale in sales:
        day = sale.created_at.date().isoformat()
        if day not in by_day_data:
            by_day_data[day] = {"sales": 0, "cogs": 0, "expenses": 0, "losses": 0}
        by_day_data[day]["sales"] += sale.total
        for item in sale.items:
            product = db.query(Product).filter(Product.id == item.product_id).first()
            if product:
                by_day_data[day]["cogs"] += item.quantity * product.weighted_cost
    
    for e in expenses:
        day = e.expense_date.date().isoformat() if isinstance(e.expense_date, datetime) else str(e.expense_date)
        if day not in by_day_data:
            by_day_data[day] = {"sales": 0, "cogs": 0, "expenses": 0, "losses": 0}
        by_day_data[day]["expenses"] += e.amount
    
    for l in losses:
        day = l.created_at.date().isoformat()
        if day not in by_day_data:
            by_day_data[day] = {"sales": 0, "cogs": 0, "expenses": 0, "losses": 0}
        by_day_data[day]["losses"] += l.total_value
    
    for day in sorted(by_day_data.keys()):
        d = by_day_data[day]
        gp = d["sales"] - d["cogs"]
        np_ = gp - d["expenses"] - d["losses"]
        ws2.append([day, d["sales"], round(d["cogs"], 2), d["expenses"], d["losses"], round(gp, 2), round(np_, 2)])
    
    # Totals row
    ws2.append(["TOTAL", total_sales, round(total_cogs, 2), total_expenses, total_losses, round(gross_profit, 2), round(net_profit, 2)])
    last_row = ws2.max_row
    for col in range(1, 8):
        ws2.cell(row=last_row, column=col).font = Font(bold=True)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"rentabilidad_{start_date}_{end_date}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/purchases", response_model=PurchasesReportResponse)
async def get_purchases_report(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    location_id: Optional[int] = None,
    product_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(RoleType.SUPERUSER, RoleType.ADMIN))
):
    """Get purchases report from stock movements of type purchase."""
    location_id = scoped_location_id(current_user, location_id)
    tenant_loc_ids = _get_tenant_location_ids(db, current_user)
    if not start_date:
        start_date = date.today() - timedelta(days=30)
    if not end_date:
        end_date = date.today()

    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())

    location_name = None
    if location_id:
        location = db.query(Location).filter(Location.id == location_id).first()
        location_name = location.name if location else None

    query = db.query(StockMovement).filter(
        StockMovement.movement_type == MovementType.PURCHASE,
        StockMovement.created_at >= start_dt,
        StockMovement.created_at <= end_dt
    )
    if tenant_loc_ids:
        query = query.filter(StockMovement.location_id.in_(tenant_loc_ids))
    if location_id:
        query = query.filter(StockMovement.location_id == location_id)
    if product_id:
        query = query.filter(StockMovement.product_id == product_id)

    movements = query.order_by(StockMovement.created_at.desc()).all()

    total_quantity = 0.0
    total_cost = 0.0
    purchases = []

    for m in movements:
        product = db.query(Product).filter(Product.id == m.product_id).first()
        loc = db.query(Location).filter(Location.id == m.location_id).first()
        user = db.query(User).filter(User.id == m.created_by_id).first() if m.created_by_id else None
        unit_cost = m.unit_cost or 0
        item_total = m.quantity * unit_cost
        total_quantity += m.quantity
        total_cost += item_total

        purchases.append(PurchaseDetail(
            id=m.id,
            date=m.created_at.strftime("%Y-%m-%d %H:%M"),
            product_code=product.code if product else "N/A",
            product_name=product.name if product else "N/A",
            location_name=loc.name if loc else "N/A",
            quantity=m.quantity,
            unit_cost=unit_cost,
            total_cost=round(item_total, 2),
            registered_by=user.full_name if user else "Sistema",
            notes=m.notes
        ))

    return PurchasesReportResponse(
        start_date=start_date,
        end_date=end_date,
        location_name=location_name or "Todas las ubicaciones",
        total_purchases=len(purchases),
        total_quantity=round(total_quantity, 2),
        total_cost=round(total_cost, 2),
        purchases=purchases
    )


@router.get("/export/purchases/excel")
async def export_purchases_excel(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    location_id: Optional[int] = None,
    product_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(RoleType.SUPERUSER, RoleType.ADMIN))
):
    from openpyxl import Workbook
    from openpyxl.styles import Font

    location_id = scoped_location_id(current_user, location_id)
    tenant_loc_ids = _get_tenant_location_ids(db, current_user)

    if not start_date:
        start_date = date.today() - timedelta(days=30)
    if not end_date:
        end_date = date.today()

    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())

    query = db.query(StockMovement).filter(
        StockMovement.movement_type == MovementType.PURCHASE,
        StockMovement.created_at >= start_dt,
        StockMovement.created_at <= end_dt
    )
    if tenant_loc_ids:
        query = query.filter(StockMovement.location_id.in_(tenant_loc_ids))
    if location_id:
        query = query.filter(StockMovement.location_id == location_id)
    if product_id:
        query = query.filter(StockMovement.product_id == product_id)

    movements = query.order_by(StockMovement.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"

    ws.append([f"Reporte de Compras - {start_date} a {end_date}"])
    ws.append([])

    headers = ["#", "Fecha", "Codigo", "Producto", "Ubicacion", "Cantidad",
               "Costo Unit.", "Costo Total", "Registrado por", "Notas"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=3, column=col).font = Font(bold=True)

    row_num = 0
    grand_qty = 0.0
    grand_cost = 0.0
    for m in movements:
        product = db.query(Product).filter(Product.id == m.product_id).first()
        loc = db.query(Location).filter(Location.id == m.location_id).first()
        user = db.query(User).filter(User.id == m.created_by_id).first() if m.created_by_id else None
        unit_cost = m.unit_cost or 0
        item_total = m.quantity * unit_cost
        grand_qty += m.quantity
        grand_cost += item_total
        row_num += 1
        ws.append([
            row_num,
            m.created_at.strftime("%Y-%m-%d %H:%M"),
            product.code if product else "N/A",
            product.name if product else "N/A",
            loc.name if loc else "N/A",
            m.quantity,
            unit_cost,
            round(item_total, 2),
            user.full_name if user else "Sistema",
            m.notes or ""
        ])

    # Totals row
    ws.append([])
    ws.append(["TOTAL", "", "", "", "", round(grand_qty, 2), "", round(grand_cost, 2)])
    last_row = ws.max_row
    for col in range(1, len(headers) + 1):
        ws.cell(row=last_row, column=col).font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"compras_{start_date}_{end_date}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/employees-summary", response_model=EmployeesReportResponse)
async def get_employees_summary_report(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    user_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(RoleType.SUPERUSER, RoleType.ADMIN))
):
    """Get performance summary for all employees in a date range."""
    if not start_date:
        start_date = date.today() - timedelta(days=30)
    if not end_date:
        end_date = date.today()
    
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())
    
    # Get all users for this tenant
    query = db.query(User)
    if current_user.tenant_id:
        query = query.filter(User.tenant_id == current_user.tenant_id)
    if user_id:
        query = query.filter(User.id == user_id)
    users = query.all()
    
    employees = []
    total_hours_all = 0.0
    total_sales_all = 0.0
    total_transactions_all = 0
    
    for user in users:
        # Get shifts in date range
        shifts = db.query(Shift).filter(
            Shift.user_id == user.id,
            Shift.start_time >= start_dt,
            Shift.start_time <= end_dt
        ).all()
        
        total_hours = 0.0
        total_sales = 0.0
        total_transactions = 0
        
        for shift in shifts:
            if shift.end_time:
                delta = shift.end_time - shift.start_time
                total_hours += delta.total_seconds() / 3600
            total_sales += shift.total_sales
            sales_count = db.query(Sale).filter(Sale.shift_id == shift.id).count()
            total_transactions += sales_count
        
        num_shifts = len(shifts)
        role_name = user.role.name if user.role else "Sin rol"
        
        employees.append(EmployeeSummary(
            user_id=user.id,
            full_name=user.full_name,
            username=user.username,
            role=role_name,
            total_hours=round(total_hours, 2),
            total_shifts=num_shifts,
            total_sales=total_sales,
            total_transactions=total_transactions,
            avg_sales_per_shift=round(total_sales / num_shifts, 2) if num_shifts > 0 else 0,
            avg_hours_per_shift=round(total_hours / num_shifts, 2) if num_shifts > 0 else 0,
            points=user.points,
            is_active=user.is_active
        ))
        
        total_hours_all += total_hours
        total_sales_all += total_sales
        total_transactions_all += total_transactions
    
    # Sort by total sales descending
    employees.sort(key=lambda e: e.total_sales, reverse=True)
    
    return EmployeesReportResponse(
        start_date=start_date,
        end_date=end_date,
        total_employees=len(employees),
        total_hours_all=round(total_hours_all, 2),
        total_sales_all=total_sales_all,
        total_transactions_all=total_transactions_all,
        employees=employees
    )


@router.get("/profitability", response_model=ProfitabilityReportResponse)
async def get_profitability_report(
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    location_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_role(RoleType.SUPERUSER, RoleType.ADMIN))
):
    """Get profitability report: Sales - Cost of Goods - Expenses - Losses = Net Profit."""
    location_id = scoped_location_id(current_user, location_id)
    tenant_loc_ids = _get_tenant_location_ids(db, current_user)
    if not start_date:
        start_date = date.today() - timedelta(days=30)
    if not end_date:
        end_date = date.today()
    
    start_dt = datetime.combine(start_date, datetime.min.time())
    end_dt = datetime.combine(end_date, datetime.max.time())
    
    location_name = None
    if location_id:
        location = db.query(Location).filter(Location.id == location_id).first()
        location_name = location.name if location else None
    
    # --- SALES ---
    sales_query = db.query(Sale).filter(
        Sale.created_at >= start_dt,
        Sale.created_at <= end_dt
    )
    if tenant_loc_ids:
        sales_query = sales_query.filter(Sale.location_id.in_(tenant_loc_ids))
    if location_id:
        sales_query = sales_query.filter(Sale.location_id == location_id)
    sales = sales_query.all()
    
    total_sales = sum(s.total for s in sales)
    total_transactions = len(sales)
    
    # --- COST OF GOODS SOLD (COGS) ---
    total_cogs = 0.0
    for sale in sales:
        for item in sale.items:
            product = db.query(Product).filter(Product.id == item.product_id).first()
            if product:
                total_cogs += item.quantity * product.weighted_cost
    
    # --- EXPENSES ---
    expenses_query = db.query(Expense).filter(
        Expense.expense_date >= start_dt,
        Expense.expense_date <= end_dt
    )
    if tenant_loc_ids:
        expenses_query = expenses_query.filter(Expense.location_id.in_(tenant_loc_ids))
    if location_id:
        expenses_query = expenses_query.filter(Expense.location_id == location_id)
    expenses = expenses_query.all()
    
    total_expenses = sum(e.amount for e in expenses)
    expenses_by_category = {}
    for e in expenses:
        cat = e.category.value if hasattr(e.category, 'value') else str(e.category)
        cat_labels = {
            'purchase': 'Compras', 'utilities': 'Servicios', 'rent': 'Arriendo',
            'salary': 'Salarios', 'maintenance': 'Mantenimiento', 'supplies': 'Insumos', 'other': 'Otros'
        }
        label = cat_labels.get(cat, cat)
        expenses_by_category[label] = expenses_by_category.get(label, 0) + e.amount
    
    # --- LOSSES ---
    losses_query = db.query(Loss).filter(
        Loss.created_at >= start_dt,
        Loss.created_at <= end_dt
    )
    if tenant_loc_ids:
        losses_query = losses_query.filter(Loss.location_id.in_(tenant_loc_ids))
    if location_id:
        losses_query = losses_query.filter(Loss.location_id == location_id)
    losses = losses_query.all()
    
    total_losses = sum(l.total_value for l in losses)
    losses_by_type = {}
    for l in losses:
        lt = l.loss_type.value if hasattr(l.loss_type, 'value') else str(l.loss_type)
        type_labels = {
            'breakage': 'Rotura', 'expiration': 'Vencimiento', 'theft': 'Robo',
            'damage': 'Daño', 'other': 'Otros'
        }
        label = type_labels.get(lt, lt)
        losses_by_type[label] = losses_by_type.get(label, 0) + l.total_value
    
    # --- CALCULATIONS ---
    gross_profit = total_sales - total_cogs
    net_profit = gross_profit - total_expenses - total_losses
    gross_margin_pct = (gross_profit / total_sales * 100) if total_sales > 0 else 0
    net_margin_pct = (net_profit / total_sales * 100) if total_sales > 0 else 0
    
    # --- BY DAY ---
    by_day_data = {}
    for sale in sales:
        day = sale.created_at.date().isoformat()
        if day not in by_day_data:
            by_day_data[day] = {"sales": 0, "cogs": 0, "expenses": 0, "losses": 0}
        by_day_data[day]["sales"] += sale.total
        for item in sale.items:
            product = db.query(Product).filter(Product.id == item.product_id).first()
            if product:
                by_day_data[day]["cogs"] += item.quantity * product.weighted_cost
    
    for e in expenses:
        day = e.expense_date.date().isoformat() if isinstance(e.expense_date, datetime) else str(e.expense_date)
        if day not in by_day_data:
            by_day_data[day] = {"sales": 0, "cogs": 0, "expenses": 0, "losses": 0}
        by_day_data[day]["expenses"] += e.amount
    
    for l in losses:
        day = l.created_at.date().isoformat()
        if day not in by_day_data:
            by_day_data[day] = {"sales": 0, "cogs": 0, "expenses": 0, "losses": 0}
        by_day_data[day]["losses"] += l.total_value
    
    by_day = []
    for day in sorted(by_day_data.keys()):
        d = by_day_data[day]
        gp = d["sales"] - d["cogs"]
        np_ = gp - d["expenses"] - d["losses"]
        by_day.append(ProfitabilityByDay(
            date=day,
            sales=d["sales"],
            cost_of_goods=d["cogs"],
            expenses=d["expenses"],
            losses=d["losses"],
            gross_profit=gp,
            net_profit=np_
        ))
    
    return ProfitabilityReportResponse(
        start_date=start_date,
        end_date=end_date,
        location_name=location_name or "Todas las ubicaciones",
        summary=ProfitabilitySummary(
            total_sales=total_sales,
            total_cost_of_goods=round(total_cogs, 2),
            gross_profit=round(gross_profit, 2),
            gross_margin_pct=round(gross_margin_pct, 1),
            total_expenses=total_expenses,
            total_losses=total_losses,
            net_profit=round(net_profit, 2),
            net_margin_pct=round(net_margin_pct, 1),
            total_transactions=total_transactions,
            expenses_by_category=expenses_by_category,
            losses_by_type=losses_by_type
        ),
        by_day=by_day
    )


@router.get("/dashboard")
async def get_dashboard_data(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    tenant_loc_ids = _get_tenant_location_ids(db, current_user)
    today = date.today()
    start_of_day = datetime.combine(today, datetime.min.time())
    end_of_day = datetime.combine(today, datetime.max.time())
    
    sales_q = db.query(func.sum(Sale.total)).filter(
        Sale.created_at >= start_of_day,
        Sale.created_at <= end_of_day
    )
    if tenant_loc_ids:
        sales_q = sales_q.filter(Sale.location_id.in_(tenant_loc_ids))
    today_sales = sales_q.scalar() or 0
    
    tx_q = db.query(Sale).filter(
        Sale.created_at >= start_of_day,
        Sale.created_at <= end_of_day
    )
    if tenant_loc_ids:
        tx_q = tx_q.filter(Sale.location_id.in_(tenant_loc_ids))
    today_transactions = tx_q.count()
    
    start_of_month = today.replace(day=1)
    month_q = db.query(func.sum(Sale.total)).filter(
        Sale.created_at >= datetime.combine(start_of_month, datetime.min.time()),
        Sale.created_at <= end_of_day
    )
    if tenant_loc_ids:
        month_q = month_q.filter(Sale.location_id.in_(tenant_loc_ids))
    month_sales = month_q.scalar() or 0
    
    low_stock_count = 0
    stock_q = db.query(ProductStock).join(Product)
    if tenant_loc_ids:
        stock_q = stock_q.filter(ProductStock.location_id.in_(tenant_loc_ids))
    stocks = stock_q.all()
    for stock in stocks:
        if stock.quantity <= stock.product.min_stock:
            low_stock_count += 1
    
    shift_q = db.query(Shift).filter(Shift.status == "open")
    if tenant_loc_ids:
        shift_q = shift_q.filter(Shift.location_id.in_(tenant_loc_ids))
    open_shifts = shift_q.count()
    
    loss_q = db.query(func.sum(Loss.total_value)).filter(
        Loss.created_at >= start_of_day,
        Loss.created_at <= end_of_day
    )
    if tenant_loc_ids:
        loss_q = loss_q.filter(Loss.location_id.in_(tenant_loc_ids))
    today_losses = loss_q.scalar() or 0
    
    expense_q = db.query(func.sum(Expense.amount)).filter(
        Expense.expense_date >= start_of_day,
        Expense.expense_date <= end_of_day
    )
    if tenant_loc_ids:
        expense_q = expense_q.filter(Expense.location_id.in_(tenant_loc_ids))
    today_expenses = expense_q.scalar() or 0
    
    return {
        "today_sales": today_sales,
        "today_transactions": today_transactions,
        "month_sales": month_sales,
        "low_stock_alerts": low_stock_count,
        "open_shifts": open_shifts,
        "today_losses": today_losses,
        "today_expenses": today_expenses
    }
